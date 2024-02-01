import os
from openpyxl import Workbook
import csv

accounting_folder = os.path.expanduser("~/Documents/accounting_files")


class FileProcessor:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.files = []
        self.base_file_names = []
        self.full_file_paths = []
        self.search_results = []
        self.all_csv_content = []
        self.list_of_csv_files = []
        self.one_file_search = []
        self.button_content = []
        self.file_dict = {}

    def get_files(self):
        # gets the files in a folder
        self.files = os.listdir(self.folder_path)
        return self.files
    
    def get_names_without_ext(self):
        self.get_files()
        for files in self.files:
            if files.endswith(".csv"):
                self.base_file_names.append(os.path.basename(files.strip(".csv")))
        return self.base_file_names

    def full_path(self):
        # appends the file names to the folder path
        for file in self.files:
            self.full_file_paths.append(os.path.join(self.folder_path, file))
        return self.full_file_paths
    
    def file_scanner(self):

        my_folder = os.listdir(self.folder_path)
        for item in my_folder:
            if item.endswith(".csv"):
                full_path = os.path.join(accounting_folder, item)
                file_name_no_extension = os.path.splitext(item)[0]

                with open(full_path, "r") as file:
                    file_content = file.read()
                    self.file_dict[file_name_no_extension] = file_content

        return self.file_dict 

    def combine_all_to_csv(self):
        # checks if files are .csv then combines them into a list
        self.get_files()
        csv_files = self.full_path()
        for file in csv_files:
            if file.endswith(".csv"):
                with open(file, "r") as acct_files:
                    for line in acct_files:
                        self.all_csv_content.append(line)
        return self.all_csv_content
    
    def list_of_csvs(self):
        self.get_files()
        csv_files = self.full_path()
        for file_name in csv_files:
            if file_name.endswith(".csv"):
                with open(file_name, "r") as files:
                    # readlines reads the entire files contents to turn into 1 item
                    content = files.readlines()
                    self.list_of_csv_files.append(content)
        print(self.list_of_csv_files)
        return self.list_of_csv_files
                    
    def search_all_files(self, search):
        self.combine_all_to_csv()
        
        for lines in self.all_csv_content:
            if search.lower() in lines.lower():
                self.search_results.append(lines)
        print(f"There are {len(self.search_results)} occurances of {search} in this file")
        for items in self.search_results:
            print(items)
        return self.search_results
    
    
    def search_one_file(self, search):
        self.one_file_search = []  # Initialize an empty list to store matching lines
        for line in self.button_content:
            if search.lower() in line.lower():
                self.one_file_search.append(line)
        print(f"There are {len(self.one_file_search)} occurrences of '{search}' in this file")


    def search_to_excel(self):
        the_workbook = Workbook()
        the_workbook.remove(the_workbook.active)  # Remove the default sheet

        keep_going = True
        while keep_going:
            search = input("Enter a search query (or type 'exit' to quit): ")
            if search.lower() == 'exit':
                keep_going = False

            self.search_all_files(search)

            if search in the_workbook.sheetnames:
                sheet = the_workbook[search]
            else:
                sheet = the_workbook.create_sheet(title=search)

            for row in self.search_results:
                split_row = [element.strip() for element in row.split(',')]
                sheet.append(split_row)
            self.search_results.clear()  # Clear the search results for the next iteration

        save_path = os.path.join(self.folder_path, "Searched_results.xlsx")
        the_workbook.save(save_path)
        print(f"Workbook saved to {save_path}")


    def combine_csvs_to_excel(self):
    # Step 1: Create a new Excel workbook
        the_workbook = Workbook()
        active_sheet = the_workbook.active
        self.full_path()

        # Step 2: Iterate over the CSV files
        for file in self.full_file_paths:
            if file.endswith(".csv"):
                # Create a new sheet for each CSV file
                sheet_name = os.path.basename(file).replace('.csv', '')
                if sheet_name == active_sheet.title:
                    sheet = active_sheet
                else:
                    sheet = the_workbook.create_sheet(title=sheet_name)
                with open(file, "r") as csv_file:
                    csv_reader = csv.reader(csv_file)
                    for row in csv_reader:
                        sheet.append(row)


        the_workbook.save(filename=os.path.join(self.folder_path, "combined_csv_files.xlsx"))

my_files = FileProcessor(accounting_folder)

